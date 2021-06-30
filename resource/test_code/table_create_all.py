import openpyxl
import datetime
#from openpyxl.styles import Font,Alignment
res_path='D:/project/qq_rob/pcr_robot/resource/'
#获取记录
 #获得成员名称
members=open(res_path+'members.txt','r',encoding='utf-8')
members_list=members.readlines()
members.close()


#伤害记录与名字的字典
attack_record={}
for name in members_list:
    name=name.strip()
    names=name.split(' ')
    attack_record[names[0]] = [0,0,0,0,0]

for i in range(0,6):
    #统计伤害
    yestoday = (datetime.date.today()-datetime.timedelta(days=i)).strftime('%Y-%m-%d')
    print(yestoday)
    record=open(res_path+'record/'+yestoday+'.txt','r',encoding='utf-8')
    record_list=record.readlines()
    record.close()
    for info in record_list:
        info_list=info.split(' ')
        attack_record[info_list[0]][int(info_list[1])-1] += int(info_list[2])


#载入basic表格
sSourceFile=res_path+'table/basic.xlsx'

sTargetFile=res_path+'table/alltime4.xlsx'
wb = openpyxl.load_workbook(sSourceFile)

#获取sheet
ws = wb['Sheet1']

#选择 5-34行 的 C列-H列 插入数据
for i in range(5,35):
    name=members_list[i-5].strip()
    names=name.split(' ')
    name_now=names[0]

    ws['b'+str(i)] = name_now
    for row in range(ord('c'),ord('g')+1):
        ws[chr(row)+str(i)]=attack_record[name_now][row-ord('c')]


wb.save(sTargetFile)
    
print("It is over")
print('hello world')